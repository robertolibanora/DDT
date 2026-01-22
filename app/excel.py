"""
Gestione sicura e thread-safe del file Excel
Operazioni atomiche con gestione errori robusta
"""
import os
import logging
import threading
from pathlib import Path
from typing import List, Dict, Any, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from contextlib import contextmanager

# NOTA: Usa get_excel_file() e get_excel_dir() da app.paths invece di EXCEL_FILE/EXCEL_DIR
# Manteniamo import per compatibilità ma useremo paths quando possibile
from app.config import EXCEL_FILE, EXCEL_DIR
from app.models import DDTData

logger = logging.getLogger(__name__)

# Lock per operazioni thread-safe
_excel_lock = threading.Lock()

# Cache per read_excel_as_dict (evita riletture continue)
_excel_cache: Optional[Dict[str, List[Dict[str, Any]]]] = None
_excel_cache_timestamp: Optional[float] = None
_excel_cache_lock = threading.Lock()

HEADERS = ["data", "mittente", "destinatario", "numero_documento", "totale_kg"]


def _ensure_excel_exists() -> None:
    """
    Crea il file Excel con gli header se non esiste
    
    IMPORTANTE: NON maschera OSError su path critici (excel directory).
    Se la directory non è scrivibile, OSError viene propagato esplicitamente.
    
    Raises:
        OSError: Se la directory excel non è scrivibile o non può essere creata
        IOError: Se c'è un errore di I/O durante la scrittura del file
    """
    from app.paths import get_excel_file
    excel_file = get_excel_file()
    
    if excel_file.exists():
        return
    
    # Assicura che la directory Excel esista usando sistema paths
    # Se la directory non è scrivibile, ensure_dir() solleverà OSError esplicitamente
    from app.paths import get_excel_dir
    excel_dir = get_excel_dir()  # Può sollevare OSError se directory non scrivibile
    
    # Crea il file Excel
    try:
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)
        
        # Formatta l'header in grassetto
        from openpyxl.styles import Font
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        wb.save(str(excel_file))
        logger.info("File Excel creato: %s", str(excel_file))
    except (OSError, IOError, PermissionError) as e:
        # Errori di I/O su file critico: propaga esplicitamente
        logger.error("Errore creazione file Excel: %s", str(e))
        raise IOError(f"Impossibile creare file Excel in {excel_dir}: {e}") from e
    except Exception as e:
        # Altri errori: propaga comunque
        logger.error("Errore creazione file Excel: %s", str(e))
        raise


@contextmanager
def _excel_operation():
    """Context manager per operazioni Excel thread-safe"""
    _excel_lock.acquire()
    try:
        _ensure_excel_exists()
        yield
    finally:
        _excel_lock.release()


def append_to_excel(data: Dict[str, Any]) -> None:
    """
    Aggiunge una riga al file Excel in modo thread-safe e atomico
    
    IMPORTANTE: NON maschera OSError/IOError su path critici (excel directory).
    Se la directory non è scrivibile, OSError viene propagato esplicitamente.
    
    Args:
        data: Dizionario con i dati del DDT (può essere dict o DDTData)
        
    Raises:
        ValueError: Se i dati non sono validi
        OSError: Se la directory excel non è scrivibile
        IOError: Se c'è un errore di I/O con il file
    """
    try:
        # Valida i dati usando Pydantic se non lo sono già
        if not isinstance(data, DDTData):
            ddt_data = DDTData(**data)
        else:
            ddt_data = data
        
        # Prepara la riga da aggiungere
        row = [
            ddt_data.data,
            ddt_data.mittente,
            ddt_data.destinatario,
            ddt_data.numero_documento,
            ddt_data.totale_kg,
        ]
        
        # Operazione atomica con lock
        # _excel_operation() chiama _ensure_excel_exists() che può sollevare OSError
        with _excel_operation():
            # Carica il workbook
            try:
                from app.paths import get_excel_file
                wb = load_workbook(str(get_excel_file()))
                ws = wb.active
            except (InvalidFileException, FileNotFoundError) as e:
                logger.error("Errore caricamento Excel: %s", str(e))
                # Ricrea il file (può sollevare OSError se directory non scrivibile)
                _ensure_excel_exists()
                from app.paths import get_excel_file
                wb = load_workbook(str(get_excel_file()))
                ws = wb.active
            
            # Verifica che l'header sia presente
            if ws.max_row == 0 or not any(ws.cell(1, col).value == HEADERS[col-1] for col in range(1, len(HEADERS)+1)):
                ws.append(HEADERS)
            
            # Aggiungi la riga
            ws.append(row)
            
            # Salva in modo sicuro
            try:
                from app.paths import get_excel_file
                excel_file = get_excel_file()
                wb.save(str(excel_file))
                logger.info("DDT aggiunto a Excel: %s", ddt_data.numero_documento)
                # Invalida cache dopo scrittura
                _invalidate_excel_cache()
            except PermissionError as e:
                logger.error("Errore: file Excel è aperto da un altro programma")
                raise IOError("Il file Excel è aperto. Chiudilo e riprova.") from e
            except (OSError, IOError) as e:
                # Errori di I/O: propaga esplicitamente
                logger.error("Errore salvataggio Excel: %s", str(e))
                raise
            except Exception as e:
                logger.error("Errore salvataggio Excel: %s", str(e))
                raise IOError(f"Errore salvataggio Excel: {e}") from e
        
    except (OSError, IOError, PermissionError):
        # Errori di I/O su path critici: propaga esplicitamente senza mascherare
        raise
    except ValueError:
        # Errori di validazione: propaga esplicitamente
        raise
    except Exception as e:
        logger.error("Errore aggiunta DDT a Excel: %s", str(e), exc_info=True)
        raise ValueError(f"Errore durante il salvataggio: {str(e)}") from e


def update_or_append_to_excel(data: Dict[str, Any]) -> bool:
    """
    Aggiorna una riga esistente in Excel se il documento esiste già, altrimenti aggiunge una nuova riga
    Identifica i documenti duplicati basandosi su numero_documento e mittente
    
    IMPORTANTE: NON maschera OSError/IOError su path critici (excel directory).
    Se la directory non è scrivibile, OSError viene propagato esplicitamente.
    
    Args:
        data: Dizionario con i dati del DDT (può essere dict o DDTData)
        
    Returns:
        True se è stato aggiornato un documento esistente, False se è stato aggiunto uno nuovo
        
    Raises:
        ValueError: Se i dati non sono validi
        OSError: Se la directory excel non è scrivibile
        IOError: Se c'è un errore di I/O con il file
    """
    try:
        # Valida i dati usando Pydantic se non lo sono già
        if not isinstance(data, DDTData):
            ddt_data = DDTData(**data)
        else:
            ddt_data = data
        
        # Prepara la riga da aggiungere/aggiornare
        row = [
            ddt_data.data,
            ddt_data.mittente,
            ddt_data.destinatario,
            ddt_data.numero_documento,
            ddt_data.totale_kg,
        ]
        
        # Normalizza per il matching
        numero_doc_norm = str(ddt_data.numero_documento).strip()
        mittente_norm = str(ddt_data.mittente).strip()
        
        # Operazione atomica con lock
        with _excel_operation():
            # Carica il workbook
            try:
                from app.paths import get_excel_file
                wb = load_workbook(str(get_excel_file()))
                ws = wb.active
            except (InvalidFileException, FileNotFoundError) as e:
                logger.error(f"Errore caricamento Excel: {e}")
                # Ricrea il file
                _ensure_excel_exists()
                from app.paths import get_excel_file
                wb = load_workbook(str(get_excel_file()))
                ws = wb.active
            
            # Verifica che l'header sia presente
            if ws.max_row == 0 or not any(ws.cell(1, col).value == HEADERS[col-1] for col in range(1, len(HEADERS)+1)):
                ws.append(HEADERS)
            
            # Cerca se esiste già un documento con lo stesso numero_documento e mittente
            # Cerca dalla fine verso l'inizio per trovare il più recente
            found_row = None
            logger.info(f"🔍 Cerca documento esistente: numero='{numero_doc_norm}', mittente='{mittente_norm}'")
            logger.info(f"   Righe da verificare: {ws.max_row - 1}")
            
            # Mostra le prime 3 righe per debug
            for debug_row in range(min(ws.max_row, 4), 1, -1):
                try:
                    debug_numero = ws.cell(debug_row, 4).value
                    debug_mittente = ws.cell(debug_row, 2).value
                    logger.info(f"   DEBUG riga {debug_row}: numero='{debug_numero}', mittente='{debug_mittente}'")
                except:
                    pass
            
            for row_num in range(ws.max_row, 1, -1):  # Dalla fine verso l'inizio
                try:
                    cell_numero = ws.cell(row_num, 4).value  # Colonna numero_documento
                    cell_mittente = ws.cell(row_num, 2).value  # Colonna mittente
                    
                    # Normalizza per il confronto
                    if cell_numero and cell_mittente:
                        cell_numero_norm = str(cell_numero).strip()
                        cell_mittente_norm = str(cell_mittente).strip()
                        
                        # Match esatto su numero documento e mittente (case-insensitive per mittente)
                        numero_match = cell_numero_norm == numero_doc_norm
                        mittente_match = cell_mittente_norm.upper() == mittente_norm.upper()
                        
                        if numero_match and mittente_match:
                            found_row = row_num
                            logger.info(f"✅ Match trovato alla riga {row_num}!")
                            break
                        elif numero_match:
                            logger.info(f"   Riga {row_num}: numero match '{cell_numero_norm}' ma mittente diverso: '{cell_mittente_norm}' != '{mittente_norm}'")
                except Exception as e:
                    logger.debug(f"Errore verifica riga {row_num}: {e}")
                    continue
            
            if not found_row:
                logger.warning(f"⚠️ Nessun documento esistente trovato per numero='{numero_doc_norm}', mittente='{mittente_norm}' - aggiungo nuovo")
            
            # Se trovato, aggiorna la riga esistente
            if found_row:
                logger.info(f"📝 Aggiornamento DDT esistente (riga {found_row}): {ddt_data.numero_documento}")
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(found_row, col_idx).value = value
                updated = True
            else:
                # Altrimenti aggiungi una nuova riga
                logger.info(f"➕ Nuovo DDT aggiunto: {ddt_data.numero_documento}")
                ws.append(row)
                updated = False
            
            # Salva in modo sicuro
            try:
                from app.paths import get_excel_file
                wb.save(str(get_excel_file()))
                if updated:
                    logger.info("DDT aggiornato in Excel: %s", ddt_data.numero_documento)
                else:
                    logger.info("DDT aggiunto a Excel: %s", ddt_data.numero_documento)
                # Invalida cache dopo scrittura
                _invalidate_excel_cache()
            except PermissionError as e:
                logger.error("Errore: file Excel è aperto da un altro programma")
                raise IOError("Il file Excel è aperto. Chiudilo e riprova.") from e
            except (OSError, IOError) as e:
                # Errori di I/O: propaga esplicitamente
                logger.error("Errore salvataggio Excel: %s", str(e))
                raise
            except Exception as e:
                logger.error("Errore salvataggio Excel: %s", str(e))
                raise IOError(f"Errore salvataggio Excel: {e}") from e
            
            return updated
        
    except (OSError, IOError, PermissionError):
        # Errori di I/O su path critici: propaga esplicitamente senza mascherare
        raise
    except ValueError:
        # Errori di validazione: propaga esplicitamente
        raise
    except Exception as e:
        logger.error("Errore aggiornamento/aggiunta DDT a Excel: %s", str(e), exc_info=True)
        raise ValueError(f"Errore durante il salvataggio: {str(e)}") from e


def _invalidate_excel_cache() -> None:
    """Invalida la cache Excel (chiamata dopo scritture)"""
    global _excel_cache, _excel_cache_timestamp
    with _excel_cache_lock:
        _excel_cache = None
        _excel_cache_timestamp = None


def read_excel_as_dict(force_reload: bool = False) -> Dict[str, List[Dict[str, Any]]]:
    """
    Legge tutto il contenuto del file Excel e restituisce un dizionario
    Usa cache con invalidazione basata su mtime del file per ridurre I/O.
    
    IMPORTANTE: NON maschera OSError/IOError su path critici (excel directory).
    Se la directory non è scrivibile, OSError viene propagato esplicitamente.
    
    Args:
        force_reload: Se True, forza ricaricamento ignorando cache
    
    Returns:
        Dizionario con chiave 'rows' contenente lista di righe
        
    Raises:
        OSError: Se la directory excel non è scrivibile
        IOError: Se c'è un errore di I/O con il file
        
    Note:
        Thread-safe, usa cache con invalidazione su mtime
    """
    global _excel_cache, _excel_cache_timestamp
    
    # _ensure_excel_exists() può sollevare OSError se directory non scrivibile
    _ensure_excel_exists()
    
    # Verifica cache se non forzato reload
    if not force_reload:
        with _excel_cache_lock:
            if _excel_cache is not None and _excel_cache_timestamp is not None:
                try:
                    from app.paths import get_excel_file
                    excel_file = get_excel_file()
                    if excel_file.exists():
                        file_mtime = excel_file.stat().st_mtime
                        if _excel_cache_timestamp == file_mtime:
                            logger.debug("Cache Excel hit: %d righe", len(_excel_cache.get("rows", [])))
                            return _excel_cache.copy()  # Ritorna copia per thread-safety
                except Exception:
                    # Se errore controllo timestamp, ricarica
                    pass
    
    # Cache miss o invalidata: ricarica
    try:
        with _excel_operation():
            try:
                from app.paths import get_excel_file
                excel_file = get_excel_file()
                wb = load_workbook(str(excel_file), data_only=True)
                ws = wb.active
            except (InvalidFileException, FileNotFoundError) as e:
                # File non valido o non trovato: può essere normale se appena creato
                logger.warning("File Excel non leggibile: %s, restituisco lista vuota", str(e))
                result = {"rows": []}
                # Aggiorna cache anche per risultato vuoto
                with _excel_cache_lock:
                    _excel_cache = result
                    try:
                        _excel_cache_timestamp = excel_file.stat().st_mtime if excel_file.exists() else None
                    except Exception:
                        _excel_cache_timestamp = None
                return result
            
            rows = []
            
            # Leggi dalla riga 2 in poi (salta header)
            # Usa max_row per leggere dall'ultima riga verso l'alto (più efficiente)
            max_row = ws.max_row
            for row_num in range(max_row, 1, -1):  # Leggi dall'ultima riga alla seconda (invertito)
                row = [cell.value for cell in ws[row_num]]
                
                # Ignora righe completamente vuote
                if not any(cell for cell in row):
                    continue
                
                # Estrai i valori con gestione None e conversioni
                data_val = row[0] if len(row) > 0 and row[0] else None
                
                # Converti date Excel in stringhe
                if data_val and hasattr(data_val, 'strftime'):
                    data_val = data_val.strftime('%Y-%m-%d')
                elif data_val is None:
                    data_val = ""
                else:
                    data_val = str(data_val)
                
                # Prepara il dizionario della riga
                row_dict = {
                    "data": data_val,
                    "mittente": str(row[1]) if len(row) > 1 and row[1] else "",
                    "destinatario": str(row[2]) if len(row) > 2 and row[2] else "",
                    "numero_documento": str(row[3]) if len(row) > 3 and row[3] else "",
                    "totale_kg": str(row[4]) if len(row) > 4 and row[4] is not None else "0",
                }
                
                rows.append(row_dict)
            
            result = {"rows": rows}
            logger.debug("Letti %d DDT da Excel (ordinati dal più recente)", len(rows))
            
            # Aggiorna cache
            with _excel_cache_lock:
                _excel_cache = result.copy()
                try:
                    _excel_cache_timestamp = excel_file.stat().st_mtime if excel_file.exists() else None
                except Exception:
                    _excel_cache_timestamp = None
            
            return result
            
    except (OSError, IOError, PermissionError):
        # Errori di I/O su path critici: propaga esplicitamente senza mascherare
        raise
    except Exception as e:
        # Altri errori: propaga comunque come IOError
        logger.error("Errore lettura Excel: %s", str(e), exc_info=True)
        raise IOError(f"Errore lettura Excel: {e}") from e


def clear_all_ddt() -> Dict[str, Any]:
    """
    Cancella tutti i DDT dal file Excel, mantenendo solo gli header
    
    Returns:
        Dizionario con risultato dell'operazione
        
    Raises:
        IOError: Se c'è un errore di I/O con il file
    """
    _ensure_excel_exists()
    
    try:
        with _excel_operation():
            # Carica il workbook
            try:
                from app.paths import get_excel_file
                wb = load_workbook(str(get_excel_file()))
                ws = wb.active
            except (InvalidFileException, FileNotFoundError) as e:
                logger.error(f"Errore caricamento Excel: {e}")
                raise IOError(f"Errore durante il caricamento del file: {str(e)}")
            
            # Conta righe prima della cancellazione
            rows_before = max(0, ws.max_row - 1)
            
            # Elimina tutte le righe tranne la prima (header)
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
            
            # Formatta l'header se non c'è già
            from openpyxl.styles import Font
            for col in range(1, len(HEADERS) + 1):
                cell = ws.cell(1, col)
                if cell.value != HEADERS[col - 1]:
                    cell.value = HEADERS[col - 1]
                cell.font = Font(bold=True)
            
            # Salva in modo sicuro
            try:
                from app.paths import get_excel_file
                wb.save(str(get_excel_file()))
                logger.info(f"Cancellati {rows_before} DDT dal file Excel")
                # Invalida cache dopo scrittura
                _invalidate_excel_cache()
                return {
                    "success": True,
                    "rows_deleted": rows_before,
                    "message": f"Cancellati {rows_before} DDT con successo"
                }
            except PermissionError as e:
                logger.error("Errore: file Excel è aperto da un altro programma")
                raise IOError("Il file Excel è aperto. Chiudilo e riprova.") from e
            except (OSError, IOError) as e:
                # Errori di I/O: propaga esplicitamente
                logger.error("Errore salvataggio Excel: %s", str(e))
                raise
            except Exception as e:
                logger.error("Errore salvataggio Excel: %s", str(e))
                raise IOError(f"Errore durante il salvataggio: {str(e)}") from e
        
    except (OSError, IOError, PermissionError):
        # Errori di I/O su path critici: propaga esplicitamente senza mascherare
        raise
    except ValueError:
        # Errori di validazione: propaga esplicitamente
        raise
    except Exception as e:
        logger.error("Errore cancellazione DDT: %s", str(e), exc_info=True)
        raise IOError(f"Errore durante la cancellazione: {str(e)}") from e


def get_excel_stats() -> Dict[str, Any]:
    """
    Ottiene statistiche dal file Excel senza caricare tutti i dati
    
    IMPORTANTE: NON maschera OSError/IOError su path critici (excel directory).
    Se la directory non è scrivibile, OSError viene propagato esplicitamente.
    
    Returns:
        Dizionario con statistiche (totale_righe, ultima_modifica, ecc.)
        
    Raises:
        OSError: Se la directory excel non è scrivibile
        IOError: Se c'è un errore di I/O con il file
    """
    # _ensure_excel_exists() può sollevare OSError se directory non scrivibile
    _ensure_excel_exists()
    
    try:
        with _excel_operation():
            from app.paths import get_excel_file
            excel_file = get_excel_file()
            if not excel_file.exists():
                return {"total_rows": 0, "file_exists": False}
            
            wb = load_workbook(str(excel_file), data_only=True)
            ws = wb.active
            
            # Conta righe (escluso header)
            total_rows = max(0, ws.max_row - 1)
            
            return {
                "total_rows": total_rows,
                "file_exists": True,
                "last_modified": excel_file.stat().st_mtime if excel_file.exists() else None,
            }
    except (OSError, IOError, PermissionError):
        # Errori di I/O su path critici: propaga esplicitamente senza mascherare
        raise
    except Exception as e:
        # Altri errori: propaga comunque come IOError
        logger.error("Errore calcolo statistiche Excel: %s", str(e), exc_info=True)
        raise IOError(f"Errore calcolo statistiche Excel: {e}") from e
